from __future__ import annotations

import io
import json
import re
import uuid
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import fitz
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE_DIR = Path(__file__).resolve().parent.parent
APP_DIR = BASE_DIR / "app"
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
PACKING_FILE = DATA_DIR / "packing.json"

app = FastAPI(title="ASN Tool GM")
app.mount("/static", StaticFiles(directory=APP_DIR / "static"), name="static")
templates = Jinja2Templates(directory=str(APP_DIR / "templates"))


# ------------------------
# storage helpers
# ------------------------
def norm_rev(value: str | int | None) -> str:
    text = str(value or "0").strip() or "0"
    digits = re.sub(r"\D", "", text)
    return f"{int(digits or '0'):02d}"


class PackingStore:
    @staticmethod
    def ensure_seed() -> None:
        if PACKING_FILE.exists():
            return
        data = {
            "single": [
                {"id": 1, "item": "316478001", "rev": "03", "packing": 55, "note": ""},
                {"id": 2, "item": "541186001", "rev": "01", "packing": 1200, "note": ""},
                {"id": 3, "item": "541183006", "rev": "01", "packing": 156, "note": ""},
                {"id": 4, "item": "316478003", "rev": "03", "packing": 55, "note": ""},
                {"id": 5, "item": "316599021", "rev": "01", "packing": 576, "note": ""},
            ],
            "pair": [
                {"id": 1, "item1": "316599012", "rev1": "01", "item2": "316600012", "rev2": "01", "packing": 576, "note": ""}
            ],
        }
        PACKING_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def load() -> dict[str, list[dict[str, Any]]]:
        PackingStore.ensure_seed()
        return json.loads(PACKING_FILE.read_text(encoding="utf-8"))

    @staticmethod
    def save(data: dict[str, list[dict[str, Any]]]) -> None:
        PACKING_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    @staticmethod
    def next_id(rows: list[dict[str, Any]]) -> int:
        return max((int(r["id"]) for r in rows), default=0) + 1


# ------------------------
# pdf parsing helpers
# ------------------------
HEADER_RE = {
    "asn_no": re.compile(r"ASN\s*No:\s*([A-Z0-9-]+)", re.I),
    "eta": re.compile(r"ETA:\s*(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2})", re.I),
    "etd": re.compile(r"ETD:\s*([^\n]*)", re.I),
}


def classify_line(line_no: str) -> str:
    raw = re.sub(r"\s+", "", (line_no or "").upper())
    if raw.startswith("C2"):
        return "CPT"
    if raw.startswith("C1"):
        return "OP"
    if raw in {"D2JOB", "GPJOB"}:
        return "GP"
    return "OTHER"



def parse_pdf(pdf_bytes: bytes, filename: str) -> dict[str, Any]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    full_text = "\n".join(page.get_text() for page in doc)
    asn_no = HEADER_RE["asn_no"].search(full_text)
    eta = HEADER_RE["eta"].search(full_text)
    etd = HEADER_RE["etd"].search(full_text)
    asn_no_text = asn_no.group(1).strip() if asn_no else Path(filename).stem
    eta_text = eta.group(1).strip() if eta else ""
    etd_text = (etd.group(1).strip() if etd else "").replace("Issued By", "").strip()

    items: list[dict[str, Any]] = []
    for page in doc:
        words = page.get_text("words")
        seq_starts: list[tuple[float, float]] = []
        for x0, y0, x1, y1, text, *_ in words:
            if x0 < 45 and text.isdigit() and y0 > 370:
                seq_starts.append((y0, y1))
        seq_starts = sorted(seq_starts, key=lambda t: t[0])
        for idx, (start_y, _) in enumerate(seq_starts):
            end_y = seq_starts[idx + 1][0] - 0.5 if idx + 1 < len(seq_starts) else 700
            row_words = [w for w in words if start_y - 1 <= w[1] <= end_y + 6]
            row_words = sorted(row_words, key=lambda v: (round(v[1], 1), v[0]))

            def texts_in(min_x: float, max_x: float) -> list[str]:
                return [w[4] for w in row_words if min_x <= w[0] < max_x]

            po_no = " ".join(texts_in(45, 125)).strip()
            item = "".join(texts_in(125, 185)).strip()
            rev = "".join(texts_in(185, 220)).strip()
            qty_text = "".join(texts_in(220, 270)).strip()
            uom = " ".join(texts_in(270, 310)).strip()
            packing_spec = "".join(texts_in(395, 465)).replace(" ", "")
            line_no = "".join(texts_in(545, 620)).replace(" ", "")
            if not item or not qty_text:
                continue
            digits = re.sub(r"[^0-9]", "", qty_text)
            if not digits:
                continue
            items.append({
                "po_no": po_no,
                "item": item,
                "rev": norm_rev(rev),
                "quantity": int(digits),
                "uom": uom,
                "packing_spec": packing_spec,
                "line_no": line_no,
            })

    line_no = next((x["line_no"] for x in items if x.get("line_no")), "")
    return {
        "asn_no": asn_no_text,
        "eta": eta_text,
        "etd": etd_text,
        "line_no": line_no,
        "group_type": classify_line(line_no),
        "items": items,
        "filename": filename,
        "pdf_bytes": pdf_bytes,
    }


# ------------------------
# conversion helpers
# ------------------------
def nearest_eta_name(asns: list[dict[str, Any]], group: str) -> str:
    vals: list[datetime] = []
    for row in asns:
        eta = row.get("eta")
        if not eta:
            continue
        try:
            vals.append(datetime.strptime(eta, "%Y-%m-%d %H:%M"))
        except ValueError:
            pass
    if not vals:
        return f"{group}_EXPORT"
    nearest = min(vals)
    return f"{group}_{nearest.strftime('%H_%M_%d%m%Y')}"



def render_pdf_to_pngs(pdf_bytes: bytes, asn_no: str, out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    created: list[Path] = []
    for idx, page in enumerate(doc):
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        suffix = f"_{idx+1}" if len(doc) > 1 else ""
        out = out_dir / f"{asn_no}{suffix}.png"
        pix.save(out)
        created.append(out)
    return created


# ------------------------
# packing logic
# ------------------------
def build_pair_lookup(pair_rows: list[dict[str, Any]]) -> tuple[dict[tuple[str, str], str], dict[str, dict[str, Any]]]:
    key_to_pair_id: dict[tuple[str, str], str] = {}
    pair_map: dict[str, dict[str, Any]] = {}
    for pair in pair_rows:
        pair_id = f"PAIR-{pair['id']}"
        pair_map[pair_id] = {
            "item1": str(pair["item1"]).strip(),
            "rev1": norm_rev(pair["rev1"]),
            "item2": str(pair["item2"]).strip(),
            "rev2": norm_rev(pair["rev2"]),
            "packing": int(pair["packing"]),
            "note": pair.get("note", ""),
        }
        key_to_pair_id[(str(pair["item1"]).strip(), norm_rev(pair["rev1"]))] = pair_id
        key_to_pair_id[(str(pair["item2"]).strip(), norm_rev(pair["rev2"]))] = pair_id
    return key_to_pair_id, pair_map



def group_items_for_excel(asn: dict[str, Any], packing_data: dict[str, list[dict[str, Any]]]) -> list[dict[str, Any]]:
    merged: dict[tuple[str, str], dict[str, Any]] = {}
    for row in asn["items"]:
        key = (str(row["item"]).strip(), norm_rev(row["rev"]))
        merged.setdefault(key, {
            "item": key[0],
            "rev": key[1],
            "quantity": 0,
            "line_no": row.get("line_no", ""),
        })
        merged[key]["quantity"] += int(row["quantity"])

    single_lookup = {(str(r["item"]).strip(), norm_rev(r["rev"])): r for r in packing_data.get("single", [])}
    pair_key_lookup, pair_map = build_pair_lookup(packing_data.get("pair", []))

    results: list[dict[str, Any]] = []
    handled_keys: set[tuple[str, str]] = set()

    def append_result(item_row: dict[str, Any], packing: int | str, carton_even: int | str, pcs_odd: int | str,
                      row_type: str, pair_group: str = "", loose_carton_count: int = 0) -> None:
        results.append({
            "stt": len(results) + 1,
            "item": item_row["item"],
            "rev": item_row["rev"],
            "quantity": item_row["quantity"],
            "packing": packing,
            "carton_even": carton_even,
            "pcs_odd": pcs_odd,
            "line_no": item_row.get("line_no", ""),
            "row_type": row_type,
            "pair_group": pair_group,
            "loose_carton_count": loose_carton_count,
        })

    # pair first
    for pair_id, pair in pair_map.items():
        key1 = (pair["item1"], pair["rev1"])
        key2 = (pair["item2"], pair["rev2"])
        if key1 not in merged or key2 not in merged:
            continue
        if key1 in handled_keys or key2 in handled_keys:
            continue

        row1 = merged[key1]
        row2 = merged[key2]
        packing = int(pair["packing"])
        qty1 = int(row1["quantity"])
        qty2 = int(row2["quantity"])
        matched = min(qty1, qty2)
        carton_even = matched // packing
        rem1 = qty1 % packing
        rem2 = qty2 % packing
        extra_diff = abs(qty1 - qty2)
        # both items still share one loose carton whenever there is any remainder or mismatch beyond full cartons
        loose_carton_count = 1 if (rem1 > 0 or rem2 > 0 or extra_diff > 0) else 0

        append_result(row1, packing, carton_even, rem1, "pair", pair_id, loose_carton_count)
        append_result(row2, packing, carton_even, rem2, "pair", pair_id, loose_carton_count)
        handled_keys.add(key1)
        handled_keys.add(key2)

    # singles and unmatched pair items fall back to single packing / unknown
    for key in sorted(merged.keys()):
        if key in handled_keys:
            continue
        row = merged[key]
        packing_row = single_lookup.get(key)
        packing = int(packing_row["packing"]) if packing_row else ""
        if packing:
            carton_even = int(row["quantity"]) // int(packing)
            pcs_odd = int(row["quantity"]) % int(packing)
            loose = 1 if pcs_odd > 0 else 0
        else:
            carton_even = ""
            pcs_odd = int(row["quantity"])
            loose = 1 if int(row["quantity"]) > 0 else 0
        append_result(row, packing, carton_even, pcs_odd, "single", "", loose)

    for idx, row in enumerate(results, start=1):
        row["stt"] = idx
    return results


# ------------------------
# excel builder
# ------------------------
def build_excel(asns: list[dict[str, Any]], output_file: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    packing = PackingStore.load()

    fill_dark = PatternFill("solid", fgColor="203864")
    fill_title = PatternFill("solid", fgColor="DDEBF7")
    fill_info = PatternFill("solid", fgColor="EDF4FC")
    fill_alt_1 = PatternFill("solid", fgColor="FFFFFF")
    fill_alt_2 = PatternFill("solid", fgColor="F7FAFD")
    fill_total = PatternFill("solid", fgColor="E2F0D9")
    fill_warn = PatternFill("solid", fgColor="FFF2CC")
    fill_pair = PatternFill("solid", fgColor="F3E5F5")

    font_title = Font(bold=True, color="FFFFFF", size=12)
    font_header = Font(bold=True, color="FFFFFF")
    font_bold = Font(bold=True, color="1F1F1F")
    font_normal = Font(color="1F1F1F")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D9E1F2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["STT", "Item", "rev", "Quantity", "Packing", "Thùng Chẵn", "PCS lẻ", "Line No."]
    widths = {"A": 8, "B": 18, "C": 8, "D": 14, "E": 12, "F": 14, "G": 12, "H": 16}
    tab_colors = {"CPT": "5B9BD5", "OP": "70AD47", "GP": "ED7D31"}

    for sheet_name in ["CPT", "OP", "GP"]:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A5"
        ws.sheet_properties.tabColor = tab_colors[sheet_name]
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = f"ASN TOOL GM - {sheet_name}"
        c.font = font_title
        c.fill = fill_dark
        c.alignment = center
        ws.row_dimensions[1].height = 24

        rows_for_sheet = [x for x in asns if x["group_type"] == sheet_name]
        if not rows_for_sheet:
            ws.merge_cells("A3:H3")
            c = ws["A3"]
            c.value = f"Không có dữ liệu {sheet_name}"
            c.font = font_bold
            c.fill = fill_info
            c.alignment = center
            continue

        cursor = 3
        for asn in rows_for_sheet:
            excel_rows = group_items_for_excel(asn, packing)

            ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=8)
            c = ws.cell(cursor, 1, f"ASN No: {asn['asn_no']}")
            c.font = font_bold
            c.fill = fill_title
            c.alignment = left
            c.border = border
            ws.row_dimensions[cursor].height = 22
            cursor += 1

            info_values = [f"ETA: {asn.get('eta','')}", f"ETD: {asn.get('etd','')}", f"Nhóm: {sheet_name}"]
            ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=3)
            ws.merge_cells(start_row=cursor, start_column=4, end_row=cursor, end_column=6)
            ws.merge_cells(start_row=cursor, start_column=7, end_row=cursor, end_column=8)
            for idx, val in zip([1,4,7], info_values):
                cell = ws.cell(cursor, idx, val)
                cell.fill = fill_info
                cell.font = font_bold
                cell.alignment = left
            for col in range(1, 9):
                ws.cell(cursor, col).border = border
            ws.row_dimensions[cursor].height = 20
            cursor += 1

            for col_idx, head in enumerate(headers, start=1):
                cell = ws.cell(cursor, col_idx, head)
                cell.fill = fill_dark
                cell.font = font_header
                cell.alignment = center
                cell.border = border
            ws.row_dimensions[cursor].height = 24
            cursor += 1

            for idx, row in enumerate(excel_rows, start=1):
                base_fill = fill_alt_1 if idx % 2 else fill_alt_2
                if row["row_type"] == "pair":
                    base_fill = fill_pair
                values = [row["stt"], row["item"], row["rev"], row["quantity"], row["packing"], row["carton_even"], row["pcs_odd"], row["line_no"]]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(cursor, col_idx, value)
                    cell.font = font_normal
                    cell.alignment = left if col_idx == 2 else center
                    cell.border = border
                    cell.fill = base_fill
                    if col_idx in {4,5,6,7} and value != "":
                        cell.number_format = '#,##0'
                if row["packing"] == "":
                    for col_idx in [5,6,7]:
                        ws.cell(cursor, col_idx).fill = fill_warn
                ws.row_dimensions[cursor].height = 22
                cursor += 1

            total_qty = sum(int(r["quantity"]) for r in excel_rows)
            total_even = sum(int(r["carton_even"]) for r in excel_rows if isinstance(r["carton_even"], int))
            loose_single = sum(1 for r in excel_rows if r["row_type"] == "single" and int(r["pcs_odd"] or 0) > 0)
            pair_groups: set[str] = set()
            loose_pair = 0
            for row in excel_rows:
                if row["row_type"] != "pair" or not row["pair_group"] or row["pair_group"] in pair_groups:
                    continue
                if int(row["loose_carton_count"] or 0) > 0:
                    loose_pair += 1
                pair_groups.add(row["pair_group"])
            total_loose_cartons = loose_single + loose_pair

            ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=3)
            c = ws.cell(cursor, 1, "TỔNG ASN")
            c.font = font_bold
            c.fill = fill_total
            c.alignment = center
            total_map = {4: total_qty, 6: total_even, 7: total_loose_cartons, 8: asn.get("line_no", "")}
            for col in range(1, 9):
                cell = ws.cell(cursor, col)
                cell.border = border
                cell.fill = fill_total
                cell.alignment = center
                if col in total_map:
                    cell.value = total_map[col]
                    if col in {4, 6, 7}:
                        cell.number_format = '#,##0'
                        cell.font = font_bold
            ws.row_dimensions[cursor].height = 22
            cursor += 2

        ws.auto_filter.ref = f"A4:H{ws.max_row}"

    wb.save(output_file)


# ------------------------
# routes
# ------------------------
@app.get("/", response_class=HTMLResponse)
async def home(request: Request) -> HTMLResponse:
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/manifest.json")
async def manifest() -> FileResponse:
    return FileResponse(APP_DIR / "static" / "manifest.json", media_type="application/manifest+json")


@app.get("/service-worker.js")
async def sw() -> FileResponse:
    return FileResponse(APP_DIR / "static" / "service-worker.js", media_type="application/javascript")


@app.get("/api/packing")
async def get_packing() -> JSONResponse:
    return JSONResponse(PackingStore.load())


@app.post("/api/packing/single")
async def add_single(item: str = Form(...), rev: str = Form(...), packing: int = Form(...), note: str = Form("")) -> JSONResponse:
    data = PackingStore.load()
    rows = data["single"]
    rows.append({"id": PackingStore.next_id(rows), "item": item.strip(), "rev": norm_rev(rev), "packing": int(packing), "note": note.strip()})
    PackingStore.save(data)
    return JSONResponse({"ok": True, "data": data})


@app.put("/api/packing/single/{row_id}")
async def update_single(row_id: int, item: str = Form(...), rev: str = Form(...), packing: int = Form(...), note: str = Form("")) -> JSONResponse:
    data = PackingStore.load()
    row = next((x for x in data["single"] if int(x["id"]) == row_id), None)
    if not row:
        raise HTTPException(404, "Không tìm thấy dòng dữ liệu")
    row.update({"item": item.strip(), "rev": norm_rev(rev), "packing": int(packing), "note": note.strip()})
    PackingStore.save(data)
    return JSONResponse({"ok": True, "data": data})


@app.delete("/api/packing/single/{row_id}")
async def delete_single(row_id: int) -> JSONResponse:
    data = PackingStore.load()
    data["single"] = [x for x in data["single"] if int(x["id"]) != row_id]
    PackingStore.save(data)
    return JSONResponse({"ok": True, "data": data})


@app.post("/api/packing/pair")
async def add_pair(item1: str = Form(...), rev1: str = Form(...), item2: str = Form(...), rev2: str = Form(...), packing: int = Form(...), note: str = Form("")) -> JSONResponse:
    data = PackingStore.load()
    rows = data["pair"]
    rows.append({
        "id": PackingStore.next_id(rows),
        "item1": item1.strip(), "rev1": norm_rev(rev1),
        "item2": item2.strip(), "rev2": norm_rev(rev2),
        "packing": int(packing), "note": note.strip(),
    })
    PackingStore.save(data)
    return JSONResponse({"ok": True, "data": data})


@app.put("/api/packing/pair/{row_id}")
async def update_pair(row_id: int, item1: str = Form(...), rev1: str = Form(...), item2: str = Form(...), rev2: str = Form(...), packing: int = Form(...), note: str = Form("")) -> JSONResponse:
    data = PackingStore.load()
    row = next((x for x in data["pair"] if int(x["id"]) == row_id), None)
    if not row:
        raise HTTPException(404, "Không tìm thấy dòng dữ liệu")
    row.update({
        "item1": item1.strip(), "rev1": norm_rev(rev1),
        "item2": item2.strip(), "rev2": norm_rev(rev2),
        "packing": int(packing), "note": note.strip(),
    })
    PackingStore.save(data)
    return JSONResponse({"ok": True, "data": data})


@app.delete("/api/packing/pair/{row_id}")
async def delete_pair(row_id: int) -> JSONResponse:
    data = PackingStore.load()
    data["pair"] = [x for x in data["pair"] if int(x["id"]) != row_id]
    PackingStore.save(data)
    return JSONResponse({"ok": True, "data": data})


@app.post("/api/packing/import-single")
async def import_single(file: UploadFile = File(...)) -> JSONResponse:
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Chỉ nhận file Excel")
    wb = load_workbook(io.BytesIO(await file.read()))
    ws = wb.active
    data = PackingStore.load()
    rows = data["single"]
    added = 0
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or raw[0] in (None, ""):
            continue
        item, rev, packing = raw[:3]
        rows.append({"id": PackingStore.next_id(rows), "item": str(item).strip(), "rev": norm_rev(rev), "packing": int(packing), "note": ""})
        added += 1
    PackingStore.save(data)
    return JSONResponse({"ok": True, "added": added, "data": data})


@app.post("/api/packing/import-pair")
async def import_pair(file: UploadFile = File(...)) -> JSONResponse:
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Chỉ nhận file Excel")
    wb = load_workbook(io.BytesIO(await file.read()))
    ws = wb.active
    data = PackingStore.load()
    rows = data["pair"]
    added = 0
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or raw[0] in (None, ""):
            continue
        item1, rev1, item2, rev2, packing = raw[:5]
        rows.append({
            "id": PackingStore.next_id(rows),
            "item1": str(item1).strip(), "rev1": norm_rev(rev1),
            "item2": str(item2).strip(), "rev2": norm_rev(rev2),
            "packing": int(packing), "note": "",
        })
        added += 1
    PackingStore.save(data)
    return JSONResponse({"ok": True, "added": added, "data": data})


@app.post("/api/process/img")
async def process_img(files: list[UploadFile] = File(...)) -> JSONResponse:
    if len(files) > 50:
        raise HTTPException(400, "Tối đa 50 PDF mỗi lần")
    job_id = uuid.uuid4().hex[:10]
    job_dir = OUTPUT_DIR / job_id
    img_dir = job_dir / "images"
    job_dir.mkdir(parents=True, exist_ok=True)

    parsed_asns: list[dict[str, Any]] = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        pdf_bytes = await f.read()
        parsed = parse_pdf(pdf_bytes, f.filename)
        parsed_asns.append(parsed)
        render_pdf_to_pngs(pdf_bytes, parsed["asn_no"], img_dir / parsed["group_type"])

    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in parsed_asns:
        grouped[row["group_type"]].append(row)

    downloads: dict[str, str] = {}
    for group in ["CPT", "OP", "GP"]:
        rows = grouped.get(group, [])
        if not rows:
            continue
        zip_name = nearest_eta_name(rows, group) + ".zip"
        zip_path = job_dir / zip_name
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for path in sorted((img_dir / group).glob("*.png")):
                zf.write(path, arcname=path.name)
        downloads[group] = f"/api/download/{job_id}/{zip_name}"

    return JSONResponse({
        "ok": True,
        "job_id": job_id,
        "counts": {g: len(grouped.get(g, [])) for g in ["CPT", "OP", "GP"]},
        "records": [{"asn_no": x["asn_no"], "eta": x["eta"], "line_no": x["line_no"], "group_type": x["group_type"]} for x in parsed_asns],
        "downloads": downloads,
    })


@app.post("/api/process/excel")
async def process_excel(files: list[UploadFile] = File(...)) -> JSONResponse:
    if len(files) > 50:
        raise HTTPException(400, "Tối đa 50 PDF mỗi lần")
    job_id = uuid.uuid4().hex[:10]
    job_dir = OUTPUT_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)
    parsed_asns: list[dict[str, Any]] = []
    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        parsed_asns.append(parse_pdf(await f.read(), f.filename))
    excel_path = job_dir / "ASN_TOOL_GM_EXPORT.xlsx"
    build_excel(parsed_asns, excel_path)
    return JSONResponse({
        "ok": True,
        "job_id": job_id,
        "download": f"/api/download/{job_id}/{excel_path.name}",
        "records": [{"asn_no": x["asn_no"], "group_type": x["group_type"], "items": len(x["items"])} for x in parsed_asns],
    })


@app.get("/api/download/{job_id}/{filename}")
async def download(job_id: str, filename: str) -> FileResponse:
    path = OUTPUT_DIR / job_id / filename
    if not path.exists():
        raise HTTPException(404, "Không tìm thấy file")
    return FileResponse(path, filename=filename)
